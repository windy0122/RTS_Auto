import logging
import MySQLdb
import redis
import random
import requests
import json
from openpyxl import load_workbook

from tools.project_path import *
from tools.read_config import ReadConfig


class StartBefore(object):
    shop_tk = None
    shop_id = None
    # shop_customer = None

    # 登录获取店铺tk和shop_id
    # 将token写入init表中
    # @staticmethod
    def get_token(self):
        url = 'https://uatleague.round-table-union.com/api/rts/base/home/login/V141'

        header = {'content-type': 'application/json',
                  'version': '1.4.1',
                  'platform': 'iOS'
                  }

        payload = {
                        "smsCodeToken": None,
                        "standbyUniqueIdentification": "_Android10",
                        "deviceUniqueIdentification": "",
                        "password": "e10adc3949ba59abbe56e057f20f883e",
                        "loginType": "1",
                        "smsCode": None,
                        "loginCode": "13411111111",
                        "source": "",
                        "applicationId": "AjwG1Pb2_9CZbEZ3QnWIc1f7HRbPBXo6-ismtWk5Y27H"
                    }

        # urllib3.disable_warnings()
        r = requests.post(url, headers=header, data=json.dumps(payload), verify=False)
        res = r.json()
        global shop_tk
        global shop_id
        shop_tk = res['val']['tk']
        shop_id = res['val']['shopId']
        self.write_back_init(test_data_path, 'init', 6, shop_tk)
        return shop_tk, shop_id

    # 查询店铺顾客手机号
    # @staticmethod
    def get_phone_num(self):
        # 打开数据库连接
        db_uat_info = eval(ReadConfig.get_config(test_config_path, 'DB', 'db_uat_config'))
        connection_uat = MySQLdb.connect(
            host=db_uat_info['host'],
            user=db_uat_info['user'],
            passwd=db_uat_info['password'],
            port=3306,
            db=db_uat_info['database'],
            charset="utf8"
        )

        # 使用cursor()方法获取操作游标
        c = connection_uat.cursor()

        # 使用execute方法执行SQL语句
        serch_customer = 'select * from tb_member_membership where Base_Shop_Id =' + shop_id

        c.execute(serch_customer)

        # 使用 fetchone() 方法获取一条数据
        self.shop_customer = []
        data = c.fetchall()
        for i in range(len(data)):
            self.shop_customer.append(data[i][5])

        # 关闭数据库连接
        connection_uat.close()

        # 返回顾客号码列表
        return self.shop_customer

    # 随机生成顾客手机号
    @staticmethod
    def create_phone():
        prelist = ['13', '14', '15', '16', '17', '18', '19']
        new_phone_num = random.choice(prelist) + ''.join(random.choice('0123456789') for i in range(9))
        return new_phone_num

    # 验证顾客号码是否是店铺会员
    # @staticmethod
    def new_customer_phone(self):
        self.get_phone_num()
        phone_num = self.create_phone()
        if phone_num in self.shop_customer:
            self.create_phone()
        else:
            return phone_num

    # 获取init表中数据(1、顾客id   2、储值卡id   3、计次卡id    4、年卡id    5、员工id    6、店铺tk)
    @staticmethod
    def get_data_init(num):
        wb = load_workbook(test_data_path)
        sheet = wb['init']
        init_data = sheet.cell(2, num).value
        return init_data

    # # 获取新建顾客id
    # @staticmethod
    # def get_customer_id():
    #     wb = load_workbook(test_data_path)
    #     sheet = wb['init']
    #     customer_id = sheet.cell(2, 1).value
    #     return customer_id
    #
    # # 获取顾客储值卡id
    # @staticmethod
    # def get_stored_card_id():
    #     wb = load_workbook(test_data_path)
    #     sheet = wb['init']
    #     stored_card_id = sheet.cell(2, 2).value
    #     return stored_card_id
    #
    # # 获取顾客计次卡id
    # @staticmethod
    # def get_time_card_id():
    #     wb = load_workbook(test_data_path)
    #     sheet = wb['init']
    #     time_card_id = sheet.cell(2, 3).value
    #     return time_card_id
    #
    # # 获取顾客年卡id
    # @staticmethod
    # def get_year_card_id():
    #     wb = load_workbook(test_data_path)
    #     sheet = wb['init']
    #     year_card_id = sheet.cell(2, 4).value
    #     return year_card_id
    #
    # # 获取员工id
    # @staticmethod
    # def get_employee_id():
    #     wb = load_workbook(test_data_path)
    #     sheet = wb['init']
    #     employee_id = sheet.cell(2, 5).value
    #     return employee_id

    # 写入result(返回数据)，写入TestResult(pass or failed)
    @staticmethod
    def write_back(file_name, sheet_name, i, result, TestResult):
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        sheet.cell(i, 7).value = result
        sheet.cell(i, 8).value = TestResult
        wb.save(file_name)

    # 写入数据到init表中
    @staticmethod
    def write_back_init(file_name, sheet_name, i, result):
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        sheet.cell(2, i).value = result
        wb.save(file_name)

    # 将顾客卡id写入到init表中
    def write_customer_card_id(self, res):
        if isinstance(res['val'], list):
            for i in range(len(res['val'])):
                if res['val'][i]['cardType'] == '1':
                    self.write_back_init(test_data_path, 'init', 2, res['val'][i]['membershipCardId'])
                elif res['val'][i]['cardType'] == '2':
                    self.write_back_init(test_data_path, 'init', 3, res['val'][i]['membershipCardId'])
                elif res['val'][i]['cardType'] == '3':
                    self.write_back_init(test_data_path, 'init', 4, res['val'][i]['membershipCardId'])
                else:
                    logging.info('顾客没有卡')

    # 将新建的员工id写入到init表中
    def write_employee_id(self, res):
        if isinstance(res['val'], list) and len(res['val']) > 1:
            self.write_back_init(test_data_path, 'init', 5, res['val'][1]['employeeId'])


if __name__ == '__main__':
    start = StartBefore()
    # print(start.get_token())
    # print(start.get_phone_num())
    # print(start.new_customer_phone())
    print(start.get_data_init(6))







